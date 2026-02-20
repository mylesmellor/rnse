"""
loader.py — Load Excel schedule and Word document.

Produces plain Python dicts/objects; all business logic lives in
validator.py and engine.py.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from docx import Document
from docx.document import Document as DocumentType

log = logging.getLogger(__name__)

# Raw schedule type: Asset_ID → {FIELD → raw cell value (may be non-numeric)}
RawSchedule = dict[str, dict[str, object]]


def load_schedule(path: Path) -> tuple[RawSchedule, list[str]]:
    """
    Read an Excel workbook and return a raw schedule dict plus the ordered
    list of field column names (excluding Asset_ID and Asset_Name).

    Returns (raw_schedule, field_names).
    raw_schedule maps Asset_ID (str) → {field_name (str) → raw_value (object)}.
    field_names is ordered as they appear in the spreadsheet.

    Raises FileNotFoundError if the path does not exist.
    The caller (validator) is responsible for all semantic checks.
    """
    log.debug("Loading schedule from %s", path)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Return minimal sentinel if sheet is missing — validator will catch this.
    if "Schedule" not in wb.sheetnames:
        log.debug("Sheet 'Schedule' not found in workbook")
        return {}, []

    ws = wb["Schedule"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return {}, []

    # First row is the header.
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Build a case-insensitive lookup for column indices.
    header_lower = {h.lower(): i for i, h in enumerate(raw_headers)}

    asset_id_col = header_lower.get("asset_id")
    asset_name_col = header_lower.get("asset_name")

    if asset_id_col is None or asset_name_col is None:
        # Validator will report the specific missing column.
        return {}, []

    # Field columns = everything that isn't Asset_ID or Asset_Name.
    special_lower = {"asset_id", "asset_name"}
    field_names: list[str] = [
        raw_headers[i].upper()
        for i, h in enumerate(raw_headers)
        if h.lower() not in special_lower and h != ""
    ]

    raw_schedule: RawSchedule = {}
    for row in rows[1:]:
        # Skip entirely empty rows.
        if all(v is None for v in row):
            continue

        asset_id_raw = row[asset_id_col]
        asset_id = str(asset_id_raw).strip().upper() if asset_id_raw is not None else ""

        if not asset_id:
            # Validator will flag empty Asset_ID cells.
            asset_id = "__EMPTY__"

        fields: dict[str, object] = {}
        for fn in field_names:
            # Find the column index for this field name (original header).
            col_idx = header_lower.get(fn.lower())
            if col_idx is not None and col_idx < len(row):
                fields[fn] = row[col_idx]
            else:
                fields[fn] = None

        if asset_id in raw_schedule:
            # Validator will flag duplicate; store under a mangled key for now
            # so we preserve all rows for error reporting.
            mangled = f"{asset_id}__DUP__"
            raw_schedule[mangled] = fields
        else:
            raw_schedule[asset_id] = fields

    log.debug("Loaded %d asset rows, %d field columns", len(raw_schedule), len(field_names))
    return raw_schedule, field_names


def load_document(path: Path) -> DocumentType:
    """
    Open a Word .docx file and return the python-docx Document object.

    Raises FileNotFoundError if the path does not exist.
    """
    log.debug("Loading document from %s", path)
    return Document(str(path))
